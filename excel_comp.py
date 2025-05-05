import openpyxl as op

file1 = input("Enter the original file: ") 
file2 = input("Enter the file with error: ") 

try:
    wb1 = op.load_workbook(file1)
    wb2 = op.load_workbook(file2)

    sheet1 = wb1.active
    sheet2 = wb2.active

    a = sheet1.max_row
    b = sheet1.max_column
    
    print(f"Comparing files with {a} rows and {b} columns...")
    
    differences_found = False
    
    for row in range(1, a + 1):
        for col in range(1, b + 1):
            val1 = sheet1.cell(row=row, column=col).value
            val2 = sheet2.cell(row=row, column=col).value
            
            if val1 != val2:
                print(f"Difference found at row: {row}, column: {col}")
                print(f"Original value: {val1} | Error file value: {val2}")
                c=input("do you want to correct the file(yes or no):")
                if(c=="yes"):
                    sheet2.cell(row=row ,column=col).value = val1
                    print(f"Original value: {val1} | corrected value: {val2}")
    
except FileNotFoundError:
    print("Error: One or both files not found. Please check the filenames.")
except Exception as e:
    print(f"An error occurred: {str(e)}")

print("Comparison complete.")